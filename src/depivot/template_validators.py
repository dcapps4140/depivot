"""
Excel template validation for depivot tool.

Validates Excel file structure, sheet templates, and data formats using a
three-phase approach: file structure (openpyxl), sheet template (openpyxl),
and DataFrame validation (pandas).
"""
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from rich.console import Console

from depivot.exceptions import TemplateError

console = Console()


class TemplateValidator:
    """
    Excel template validation orchestrator.

    Validates Excel files in three phases:
    1. File structure validation (fast, openpyxl, no data load)
    2. Sheet template validation (openpyxl, per sheet)
    3. DataFrame validation (pandas, after data load)
    """

    def __init__(self, config: Dict):
        """
        Initialize template validator from config.

        Args:
            config: Template validation configuration dictionary with structure:
                {
                    "enabled": bool,
                    "file_structure": [check_configs],
                    "sheet_template": [check_configs],
                    "dataframe_template": [check_configs],
                    "settings": {"stop_on_error": bool, "verbose": bool}
                }
        """
        self.config = config
        self.enabled = config.get("enabled", True)
        self.settings = config.get("settings", {})
        self.stop_on_error = self.settings.get("stop_on_error", True)
        self.verbose = self.settings.get("verbose", False)

        self.file_structure_checks = config.get("file_structure", [])
        self.sheet_template_checks = config.get("sheet_template", [])
        self.dataframe_template_checks = config.get("dataframe_template", [])

    def validate_file_structure(self, input_file: Path) -> None:
        """
        Phase 1: Fast structure validation using openpyxl.

        Validates file-level structure without loading data:
        - Expected sheets exist
        - Sheet count within range
        - File integrity

        Args:
            input_file: Path to Excel file

        Raises:
            TemplateError: If validation fails with error severity
        """
        if not self.enabled:
            return

        try:
            wb = load_workbook(input_file, read_only=True, data_only=True)
        except Exception as e:
            raise TemplateError(f"Cannot open Excel file: {e}")

        for check in self.file_structure_checks:
            if not check.get("enabled", True):
                continue

            check_type = check.get("check")

            if check_type == "expected_sheets":
                self._check_expected_sheets(wb, check)
            elif check_type == "sheet_count":
                self._check_sheet_count(wb, check)

        wb.close()

    def validate_sheet_template(self, input_file: Path, sheet_name: str) -> None:
        """
        Phase 2: Sheet template validation using openpyxl.

        Validates sheet-level template without loading full data:
        - Header row content and position
        - Merged cells detection
        - Cell format validation

        Args:
            input_file: Path to Excel file
            sheet_name: Name of sheet to validate

        Raises:
            TemplateError: If validation fails with error severity
        """
        if not self.enabled:
            return

        try:
            wb = load_workbook(input_file, read_only=False, data_only=False)
        except Exception as e:
            raise TemplateError(f"Cannot open Excel file for sheet validation: {e}")

        if sheet_name not in wb.sheetnames:
            wb.close()
            return

        ws = wb[sheet_name]

        for check in self.sheet_template_checks:
            if not check.get("enabled", True):
                continue

            check_type = check.get("check")

            if check_type == "header_row":
                self._check_header_row(ws, sheet_name, check)
            elif check_type == "merged_cells":
                self._check_merged_cells(ws, sheet_name, check)
            elif check_type == "cell_formats":
                self._check_cell_formats(ws, sheet_name, check)

        wb.close()

    def validate_dataframe_template(self, df: pd.DataFrame, sheet_name: str) -> None:
        """
        Phase 3: DataFrame validation using pandas.

        Validates DataFrame structure after data is loaded:
        - Column order
        - Required columns presence
        - Sample data validation

        Args:
            df: DataFrame to validate
            sheet_name: Name of sheet (for error messages)

        Raises:
            TemplateError: If validation fails with error severity
        """
        if not self.enabled:
            return

        for check in self.dataframe_template_checks:
            if not check.get("enabled", True):
                continue

            check_type = check.get("check")

            if check_type == "column_order":
                self._check_column_order(df, sheet_name, check)
            elif check_type == "required_columns":
                self._check_required_columns(df, sheet_name, check)

    # =========================================================================
    # FILE STRUCTURE CHECKS
    # =========================================================================

    def _check_expected_sheets(self, wb, check: Dict) -> None:
        """
        Check required sheets exist in workbook.

        Params:
            required_sheets: List of required sheet names
            allow_extra_sheets: If False, fail on extra sheets
            exact_match: If True, found sheets must match exactly

        Raises:
            TemplateError: If required sheets missing or extra sheets found
        """
        required = check["params"].get("required_sheets", [])
        allow_extra = check["params"].get("allow_extra_sheets", True)
        found = wb.sheetnames

        # Check for missing required sheets
        missing = [s for s in required if s not in found]
        if missing:
            msg = check.get("message", "Missing sheets: {sheets}")
            raise TemplateError(
                msg.format(sheets=", ".join(missing)) +
                f"\n  Expected: {', '.join(required)}" +
                f"\n  Found: {', '.join(found)}"
            )

        # Check for extra sheets if not allowed
        if not allow_extra:
            extra = [s for s in found if s not in required]
            if extra:
                msg = f"Extra sheets found: {', '.join(extra)}"
                if check.get("severity") == "error":
                    raise TemplateError(msg)
                else:
                    console.print(f"[yellow]Template Warning: {msg}[/yellow]")

    def _check_sheet_count(self, wb, check: Dict) -> None:
        """
        Check sheet count is within expected range.

        Params:
            min_sheets: Minimum sheet count (inclusive)
            max_sheets: Maximum sheet count (inclusive)

        Raises:
            TemplateError: If sheet count outside range
        """
        min_sheets = check["params"].get("min_sheets", 1)
        max_sheets = check["params"].get("max_sheets", None)

        count = len(wb.sheetnames)

        if count < min_sheets:
            msg = check.get("message", "Sheet count too low: {count}")
            raise TemplateError(
                msg.format(count=count) +
                f" (minimum: {min_sheets})"
            )

        if max_sheets is not None and count > max_sheets:
            msg = check.get("message", "Sheet count too high: {count}")
            severity = check.get("severity", "warning")

            error_msg = msg.format(count=count) + f" (maximum: {max_sheets})"

            if severity == "error":
                raise TemplateError(error_msg)
            else:
                console.print(f"[yellow]Template Warning: {error_msg}[/yellow]")

    # =========================================================================
    # SHEET TEMPLATE CHECKS
    # =========================================================================

    def _check_header_row(self, ws, sheet_name: str, check: Dict) -> None:
        """
        Check header row content and position.

        Params:
            row_number: Row number (1-indexed, Excel style)
            expected_columns: List of expected column names
            exact_order: If True, columns must be in exact order
            allow_extra_columns: If True, allow columns not in expected list

        Raises:
            TemplateError: If header row doesn't match expectations
        """
        row_num = check["params"].get("row_number", 1)
        expected = check["params"].get("expected_columns", [])
        exact_order = check["params"].get("exact_order", False)
        allow_extra = check["params"].get("allow_extra_columns", True)

        # Read header row (convert 1-indexed to actual row)
        actual = []
        for cell in ws[row_num]:
            if cell.value is not None:
                actual.append(str(cell.value).strip())

        # Check for missing columns
        missing = [col for col in expected if col not in actual]
        if missing:
            msg = check.get("message", "Header row mismatch in sheet '{sheet}'")
            raise TemplateError(
                msg.format(sheet=sheet_name) +
                f"\n  Missing columns: {', '.join(missing)}" +
                f"\n  Expected: {', '.join(expected)}" +
                f"\n  Found at row {row_num}: {', '.join(actual)}"
            )

        # Check for extra columns if not allowed
        if not allow_extra:
            extra = [col for col in actual if col not in expected]
            if extra:
                error_msg = (
                    f"Extra columns in sheet '{sheet_name}': {', '.join(extra)}"
                )
                if check.get("severity") == "error":
                    raise TemplateError(error_msg)
                else:
                    console.print(f"[yellow]Template Warning: {error_msg}[/yellow]")

        # Check exact order if required
        if exact_order:
            # Filter actual to only expected columns, maintaining order
            expected_in_actual = [col for col in actual if col in expected]
            if expected_in_actual != expected:
                raise TemplateError(
                    f"Column order mismatch in sheet '{sheet_name}'" +
                    f"\n  Expected order: {', '.join(expected)}" +
                    f"\n  Found order: {', '.join(expected_in_actual)}"
                )

    def _check_merged_cells(self, ws, sheet_name: str, check: Dict) -> None:
        """
        Check for merged cells in sheet.

        Params:
            allowed: If False, fail on any merged cells
            allowed_ranges: List of allowed merged cell ranges (e.g., ["A1:B1"])

        Raises:
            TemplateError: If merged cells found and not allowed
        """
        allowed = check["params"].get("allowed", False)
        allowed_ranges = check["params"].get("allowed_ranges", [])

        merged = list(ws.merged_cells.ranges)

        if not allowed and merged:
            # Check if all merged cells are in allowed ranges
            if allowed_ranges:
                disallowed = [
                    str(r) for r in merged
                    if str(r) not in allowed_ranges
                ]
                if not disallowed:
                    return  # All merged cells are allowed

                ranges_str = ", ".join(disallowed[:5])
                if len(disallowed) > 5:
                    ranges_str += f", ... and {len(disallowed) - 5} more"
            else:
                ranges_str = ", ".join(str(r) for r in merged[:5])
                if len(merged) > 5:
                    ranges_str += f", ... and {len(merged) - 5} more"

            msg = check.get("message", "Merged cells detected in '{sheet}': {ranges}")
            error_msg = msg.format(sheet=sheet_name, ranges=ranges_str)

            if check.get("severity") == "error":
                raise TemplateError(error_msg)
            else:
                console.print(f"[yellow]Template Warning: {error_msg}[/yellow]")

    def _check_cell_formats(self, ws, sheet_name: str, check: Dict) -> None:
        """
        Check cell formats match expectations.

        Params:
            check_types: If True, check data types
            numeric_columns: List of columns that should have numeric format
            max_rows_to_check: Maximum rows to check (default 100)

        Raises:
            TemplateError: If cell formats don't match expectations
        """
        check_types = check["params"].get("check_types", True)
        numeric_columns = check["params"].get("numeric_columns", [])
        max_rows = check["params"].get("max_rows_to_check", 100)

        if not check_types or not numeric_columns:
            return

        # Get header row to find column indices
        header_row = 1
        headers = {}
        for idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx

        # Check numeric columns have numeric format
        issues = []
        for col_name in numeric_columns:
            if col_name not in headers:
                continue

            col_idx = headers[col_name]

            # Check first few rows for format
            for row_idx in range(2, min(max_rows + 2, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.value is None:
                    continue

                # Check if cell is numeric type
                if not isinstance(cell.value, (int, float)):
                    # Also check number format
                    if cell.number_format == "General" or "0" not in str(cell.number_format):
                        issues.append(f"{col_name} (row {row_idx})")
                        break  # Only report first issue per column

        if issues:
            msg = check.get("message", "Cell format issues in '{sheet}'")
            error_msg = (
                msg.format(sheet=sheet_name) +
                f"\n  Non-numeric formats in: {', '.join(issues[:5])}"
            )

            if check.get("severity") == "error":
                raise TemplateError(error_msg)
            else:
                console.print(f"[yellow]Template Warning: {error_msg}[/yellow]")

    # =========================================================================
    # DATAFRAME CHECKS
    # =========================================================================

    def _check_column_order(self, df: pd.DataFrame, sheet_name: str, check: Dict) -> None:
        """
        Check DataFrame column order matches expectations.

        Params:
            expected_order: List of column names in expected order
            strict: If True, ALL columns must match order exactly

        Raises:
            TemplateError: If column order doesn't match
        """
        expected_order = check["params"].get("expected_order", [])
        strict = check["params"].get("strict", False)

        actual = df.columns.tolist()

        if strict:
            # All columns must match exactly
            if actual != expected_order:
                msg = check.get("message", "Column order mismatch in '{sheet}'")
                error_msg = (
                    msg.format(sheet=sheet_name) +
                    f"\n  Expected: {', '.join(expected_order)}" +
                    f"\n  Found: {', '.join(actual)}"
                )

                if check.get("severity") == "error":
                    raise TemplateError(error_msg)
                else:
                    console.print(f"[yellow]Template Warning: {error_msg}[/yellow]")
        else:
            # Only expected columns must be in order (others can be anywhere)
            expected_in_actual = [col for col in actual if col in expected_order]
            if expected_in_actual != expected_order:
                msg = check.get("message", "Column order mismatch in '{sheet}'")
                error_msg = (
                    msg.format(sheet=sheet_name) +
                    f"\n  Expected order: {', '.join(expected_order)}" +
                    f"\n  Found order: {', '.join(expected_in_actual)}"
                )

                if check.get("severity") == "error":
                    raise TemplateError(error_msg)
                else:
                    console.print(f"[yellow]Template Warning: {error_msg}[/yellow]")

    def _check_required_columns(self, df: pd.DataFrame, sheet_name: str, check: Dict) -> None:
        """
        Check required columns are present in DataFrame.

        Params:
            columns: List of required column names

        Raises:
            TemplateError: If required columns missing
        """
        required = check["params"].get("columns", [])
        actual = df.columns.tolist()

        missing = [col for col in required if col not in actual]

        if missing:
            msg = check.get("message", "Required columns missing in '{sheet}': {missing}")
            raise TemplateError(
                msg.format(sheet=sheet_name, missing=", ".join(missing)) +
                f"\n  Required: {', '.join(required)}" +
                f"\n  Found: {', '.join(actual)}"
            )
