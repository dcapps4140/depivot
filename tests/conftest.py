"""Pytest configuration and shared fixtures for depivot tests."""
import tempfile
from pathlib import Path
from typing import Dict, List

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import numbers


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_dataframe():
    """Create a sample DataFrame for testing."""
    return pd.DataFrame({
        "Site": ["Site A", "Site B", "Site C", "Site D", "Site E"],
        "Category": ["Cat1", "Cat2", "Cat1", "Cat3", "Cat2"],
        "Jan": [100, 200, 150, 300, 250],
        "Feb": [110, 210, 160, 310, 260],
        "Mar": [120, 220, 170, 320, 270],
        "Apr": [130, 230, 180, 330, 280],
    })


@pytest.fixture
def sample_excel_file(temp_dir, sample_dataframe):
    """Create a sample Excel file for testing."""
    file_path = temp_dir / "test_data.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Write headers
    headers = ["Site", "Category", "Jan", "Feb", "Mar", "Apr"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, row in enumerate(sample_dataframe.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(file_path)
    return file_path


@pytest.fixture
def multi_sheet_excel_file(temp_dir, sample_dataframe):
    """Create a multi-sheet Excel file for testing."""
    file_path = temp_dir / "multi_sheet_data.xlsx"

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        sample_dataframe.to_excel(writer, sheet_name="Sheet1", index=False)
        sample_dataframe.to_excel(writer, sheet_name="Sheet2", index=False)
        sample_dataframe.to_excel(writer, sheet_name="Metadata", index=False)

    return file_path


@pytest.fixture
def excel_file_with_merged_cells(temp_dir):
    """Create an Excel file with merged cells for testing."""
    file_path = temp_dir / "merged_cells.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Merge cells B1:C1
    ws.merge_cells('B1:C1')
    ws['B1'] = "Merged Header"

    # Write headers
    ws['A1'] = "Site"
    ws['D1'] = "Jan"
    ws['E1'] = "Feb"

    # Write some data
    ws['A2'] = "Site A"
    ws['B2'] = "Value1"
    ws['C2'] = "Value2"
    ws['D2'] = 100
    ws['E2'] = 110

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_file_with_wrong_headers(temp_dir):
    """Create an Excel file with incorrect headers for testing."""
    file_path = temp_dir / "wrong_headers.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Write incorrect headers
    headers = ["Location", "Type", "January", "February"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    wb.save(file_path)
    return file_path


@pytest.fixture
def dataframe_with_nulls():
    """Create a DataFrame with NULL values for testing."""
    return pd.DataFrame({
        "Site": ["Site A", "Site B", None, "Site D", None],
        "Category": ["Cat1", None, "Cat1", "Cat3", "Cat2"],
        "Jan": [100, 200, 150, None, 250],
        "Feb": [110, None, 160, 310, 260],
    })


@pytest.fixture
def dataframe_with_duplicates():
    """Create a DataFrame with duplicate rows for testing."""
    return pd.DataFrame({
        "Site": ["Site A", "Site B", "Site A", "Site C", "Site B"],
        "Category": ["Cat1", "Cat2", "Cat1", "Cat3", "Cat2"],
        "Jan": [100, 200, 100, 300, 200],
        "Feb": [110, 210, 110, 310, 210],
    })


@pytest.fixture
def template_validation_config():
    """Create a sample template validation configuration."""
    return {
        "enabled": True,
        "file_structure": [
            {
                "check": "expected_sheets",
                "enabled": True,
                "severity": "error",
                "params": {
                    "required_sheets": ["Test Sheet"],
                    "allow_extra_sheets": True,
                },
                "message": "Required sheet not found"
            }
        ],
        "sheet_template": [
            {
                "check": "header_row",
                "enabled": True,
                "severity": "error",
                "params": {
                    "row_number": 1,
                    "expected_columns": ["Site", "Category", "Jan", "Feb", "Mar", "Apr"],
                    "exact_order": False,
                    "allow_extra_columns": True,
                },
                "message": "Header row mismatch in sheet '{sheet}'"
            }
        ],
        "dataframe_template": [
            {
                "check": "required_columns",
                "enabled": True,
                "severity": "error",
                "params": {
                    "columns": ["Site", "Category"],
                },
                "message": "Required columns missing in '{sheet}': {missing}"
            }
        ],
        "settings": {
            "stop_on_error": True,
            "verbose": False,
        }
    }


@pytest.fixture
def data_quality_config():
    """Create a sample data quality validation configuration."""
    return {
        "enabled": True,
        "pre_processing": [
            {
                "rule": "check_null_values",
                "enabled": True,
                "severity": "warning",
                "params": {
                    "columns": ["Site", "Category"],
                    "threshold": 0.2,
                },
                "message": "Excessive NULL values in {column}: {percent}%"
            }
        ],
        "post_processing": [
            {
                "rule": "check_row_count",
                "enabled": True,
                "severity": "error",
                "params": {
                    "min_ratio": 0.95,
                    "max_ratio": 1.05,
                },
                "message": "Row count mismatch"
            }
        ],
        "validation_settings": {
            "stop_on_error": True,
            "max_warnings_display": 10,
        }
    }
