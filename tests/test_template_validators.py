"""Tests for Excel template validation."""
import pytest
import pandas as pd
from pathlib import Path
from openpyxl import Workbook

from depivot.template_validators import TemplateValidator
from depivot.exceptions import TemplateError


class TestTemplateValidator:
    """Tests for TemplateValidator class."""

    def test_validator_initialization(self, template_validation_config):
        """Test TemplateValidator initializes correctly."""
        validator = TemplateValidator(template_validation_config)

        assert validator.enabled is True
        assert validator.stop_on_error is True
        assert len(validator.file_structure_checks) == 1
        assert len(validator.sheet_template_checks) == 1
        assert len(validator.dataframe_template_checks) == 1

    def test_validator_disabled(self):
        """Test validator can be disabled."""
        config = {"enabled": False}
        validator = TemplateValidator(config)

        assert validator.enabled is False

    def test_validator_empty_config(self):
        """Test validator with empty configuration."""
        config = {}
        validator = TemplateValidator(config)

        assert validator.enabled is True
        assert validator.file_structure_checks == []
        assert validator.sheet_template_checks == []
        assert validator.dataframe_template_checks == []

    def test_validator_settings(self):
        """Test validator settings extraction."""
        config = {
            "enabled": True,
            "settings": {
                "stop_on_error": False,
                "verbose": True
            }
        }
        validator = TemplateValidator(config)

        assert validator.stop_on_error is False
        assert validator.verbose is True


class TestFileStructureValidation:
    """Tests for file structure validation (Phase 1)."""

    def test_expected_sheets_pass(self, sample_excel_file, template_validation_config):
        """Test expected sheets validation passes with correct sheet."""
        validator = TemplateValidator(template_validation_config)

        # Should not raise exception
        validator.validate_file_structure(sample_excel_file)

    def test_expected_sheets_fail(self, sample_excel_file):
        """Test expected sheets validation fails with missing sheet."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "expected_sheets",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "required_sheets": ["Nonexistent Sheet"],
                        "allow_extra_sheets": True,
                    },
                    "message": "Required sheet not found"
                }
            ],
            "sheet_template": [],
            "dataframe_template": [],
            "settings": {}
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Required sheet not found"):
            validator.validate_file_structure(sample_excel_file)

    def test_expected_sheets_extra_not_allowed(self, multi_sheet_excel_file):
        """Test validation fails when extra sheets present and not allowed."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "expected_sheets",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "required_sheets": ["Sheet1"],
                        "allow_extra_sheets": False,
                    },
                    "message": "Extra sheets found"
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Extra sheets found"):
            validator.validate_file_structure(multi_sheet_excel_file)

    def test_expected_sheets_extra_warning(self, multi_sheet_excel_file):
        """Test validation warns on extra sheets with warning severity."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "expected_sheets",
                    "enabled": True,
                    "severity": "warning",
                    "params": {
                        "required_sheets": ["Sheet1"],
                        "allow_extra_sheets": False,
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise, just warn
        validator.validate_file_structure(multi_sheet_excel_file)

    def test_sheet_count_min(self, sample_excel_file):
        """Test minimum sheet count validation."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "sheet_count",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "min_sheets": 5
                    },
                    "message": "Too few sheets"
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Too few sheets"):
            validator.validate_file_structure(sample_excel_file)

    def test_sheet_count_max(self, multi_sheet_excel_file):
        """Test maximum sheet count validation."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "sheet_count",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "min_sheets": 1,
                        "max_sheets": 2
                    },
                    "message": "Too many sheets"
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Too many sheets"):
            validator.validate_file_structure(multi_sheet_excel_file)

    def test_sheet_count_max_warning(self, multi_sheet_excel_file):
        """Test maximum sheet count validation with warning severity."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "sheet_count",
                    "enabled": True,
                    "severity": "warning",
                    "params": {
                        "min_sheets": 1,
                        "max_sheets": 2
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise, just warn
        validator.validate_file_structure(multi_sheet_excel_file)

    def test_sheet_count_valid_range(self, multi_sheet_excel_file):
        """Test sheet count validation passes within valid range."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "sheet_count",
                    "enabled": True,
                    "params": {
                        "min_sheets": 2,
                        "max_sheets": 5
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise
        validator.validate_file_structure(multi_sheet_excel_file)

    def test_file_structure_disabled_check(self, sample_excel_file):
        """Test disabled checks are skipped."""
        config = {
            "enabled": True,
            "file_structure": [
                {
                    "check": "sheet_count",
                    "enabled": False,
                    "params": {
                        "min_sheets": 100
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise because check is disabled
        validator.validate_file_structure(sample_excel_file)

    def test_file_structure_invalid_file(self, temp_dir):
        """Test validation fails with invalid Excel file."""
        invalid_file = temp_dir / "invalid.xlsx"
        invalid_file.write_text("Not an Excel file")

        config = {"enabled": True, "file_structure": []}
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Cannot open Excel file"):
            validator.validate_file_structure(invalid_file)

    def test_file_structure_disabled_validator(self, sample_excel_file):
        """Test disabled validator skips validation."""
        config = {"enabled": False}
        validator = TemplateValidator(config)

        # Should not raise, validator is disabled
        validator.validate_file_structure(sample_excel_file)


class TestSheetTemplateValidation:
    """Tests for sheet template validation (Phase 2)."""

    def test_header_row_validation_pass(self, sample_excel_file, template_validation_config):
        """Test header row validation passes with correct headers."""
        validator = TemplateValidator(template_validation_config)

        # Should not raise exception
        validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_header_row_missing_columns(self, sample_excel_file):
        """Test header row validation fails with missing columns."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "header_row",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "row_number": 1,
                        "expected_columns": ["Site", "Category", "Nonexistent"]
                    },
                    "message": "Header mismatch"
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Header mismatch"):
            validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_header_row_exact_order(self, sample_excel_file):
        """Test header row validation with exact order requirement."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "header_row",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "row_number": 1,
                        "expected_columns": ["Feb", "Jan", "Mar"],  # Wrong order - Feb before Jan
                        "exact_order": True
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Column order mismatch"):
            validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_header_row_extra_columns_not_allowed(self, sample_excel_file):
        """Test header row validation fails with extra columns when not allowed."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "header_row",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "row_number": 1,
                        "expected_columns": ["Site", "Category"],
                        "allow_extra_columns": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Extra columns"):
            validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_header_row_extra_columns_warning(self, sample_excel_file):
        """Test header row validation warns on extra columns with warning severity."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "header_row",
                    "enabled": True,
                    "severity": "warning",
                    "params": {
                        "row_number": 1,
                        "expected_columns": ["Site"],
                        "allow_extra_columns": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise, just warn
        validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_merged_cells_detection(self, excel_file_with_merged_cells):
        """Test merged cells are detected."""
        config = {
            "enabled": True,
            "file_structure": [],
            "sheet_template": [
                {
                    "check": "merged_cells",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "allowed": False,
                    },
                    "message": "Merged cells detected"
                }
            ],
            "dataframe_template": [],
            "settings": {}
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Merged cells detected"):
            validator.validate_sheet_template(excel_file_with_merged_cells, "Test Sheet")

    def test_merged_cells_allowed(self, excel_file_with_merged_cells):
        """Test merged cells validation passes when allowed."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "merged_cells",
                    "enabled": True,
                    "params": {
                        "allowed": True
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise
        validator.validate_sheet_template(excel_file_with_merged_cells, "Test Sheet")

    def test_merged_cells_allowed_ranges(self, excel_file_with_merged_cells):
        """Test merged cells validation with allowed ranges."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "merged_cells",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "allowed": False,
                        "allowed_ranges": ["B1:C1"]
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise because B1:C1 is in allowed ranges
        validator.validate_sheet_template(excel_file_with_merged_cells, "Test Sheet")

    def test_merged_cells_warning(self, excel_file_with_merged_cells):
        """Test merged cells validation warns with warning severity."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "merged_cells",
                    "enabled": True,
                    "severity": "warning",
                    "params": {
                        "allowed": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise, just warn
        validator.validate_sheet_template(excel_file_with_merged_cells, "Test Sheet")

    def test_cell_formats_numeric_columns(self, temp_dir):
        """Test cell format validation for numeric columns."""
        file_path = temp_dir / "formats.xlsx"
        wb = Workbook()
        ws = wb.active

        # Write headers
        ws['A1'] = "Site"
        ws['B1'] = "Amount"

        # Write non-numeric value in Amount column
        ws['A2'] = "Site A"
        ws['B2'] = "Not a number"

        wb.save(file_path)
        wb.close()

        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "cell_formats",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "check_types": True,
                        "numeric_columns": ["Amount"],
                        "max_rows_to_check": 10
                    },
                    "message": "Format issues"
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Format issues"):
            validator.validate_sheet_template(file_path, "Sheet")

    def test_cell_formats_disabled(self, sample_excel_file):
        """Test cell format validation is skipped when disabled."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "cell_formats",
                    "enabled": True,
                    "params": {
                        "check_types": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise
        validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_cell_formats_warning(self, temp_dir):
        """Test cell format validation warns with warning severity."""
        file_path = temp_dir / "formats.xlsx"
        wb = Workbook()
        ws = wb.active

        ws['A1'] = "Site"
        ws['B1'] = "Amount"
        ws['A2'] = "Site A"
        ws['B2'] = "Text"

        wb.save(file_path)
        wb.close()

        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "cell_formats",
                    "enabled": True,
                    "severity": "warning",
                    "params": {
                        "check_types": True,
                        "numeric_columns": ["Amount"]
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise, just warn
        validator.validate_sheet_template(file_path, "Sheet")

    def test_sheet_template_missing_sheet(self, sample_excel_file):
        """Test sheet template validation skips missing sheets."""
        config = {
            "enabled": True,
            "sheet_template": [
                {
                    "check": "header_row",
                    "enabled": True,
                    "params": {"row_number": 1, "expected_columns": ["A"]}
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise for missing sheet
        validator.validate_sheet_template(sample_excel_file, "Nonexistent Sheet")

    def test_sheet_template_disabled_validator(self, sample_excel_file):
        """Test disabled validator skips sheet validation."""
        config = {"enabled": False}
        validator = TemplateValidator(config)

        # Should not raise, validator is disabled
        validator.validate_sheet_template(sample_excel_file, "Test Sheet")

    def test_sheet_template_invalid_file(self, temp_dir):
        """Test sheet validation fails with invalid file."""
        invalid_file = temp_dir / "invalid.xlsx"
        invalid_file.write_text("Not an Excel file")

        config = {"enabled": True, "sheet_template": []}
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Cannot open Excel file"):
            validator.validate_sheet_template(invalid_file, "Sheet")


class TestDataFrameValidation:
    """Tests for DataFrame validation (Phase 3)."""

    def test_required_columns_pass(self, sample_dataframe, template_validation_config):
        """Test required columns validation passes."""
        validator = TemplateValidator(template_validation_config)

        # Should not raise exception
        validator.validate_dataframe_template(sample_dataframe, "Test Sheet")

    def test_required_columns_fail(self, sample_dataframe):
        """Test required columns validation fails with missing columns."""
        config = {
            "enabled": True,
            "file_structure": [],
            "sheet_template": [],
            "dataframe_template": [
                {
                    "check": "required_columns",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "columns": ["Site", "Category", "Nonexistent"],
                    },
                    "message": "Required columns missing"
                }
            ],
            "settings": {}
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Required columns missing"):
            validator.validate_dataframe_template(sample_dataframe, "Test Sheet")

    def test_column_order_non_strict(self, sample_dataframe):
        """Test column order validation (non-strict mode)."""
        config = {
            "enabled": True,
            "dataframe_template": [
                {
                    "check": "column_order",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "expected_order": ["Site", "Category", "Jan"],
                        "strict": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise - these columns are in correct order
        validator.validate_dataframe_template(sample_dataframe, "Test Sheet")

    def test_column_order_non_strict_fail(self, sample_dataframe):
        """Test column order validation fails when order is wrong."""
        config = {
            "enabled": True,
            "dataframe_template": [
                {
                    "check": "column_order",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "expected_order": ["Category", "Site"],  # Wrong order
                        "strict": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Column order mismatch"):
            validator.validate_dataframe_template(sample_dataframe, "Test Sheet")

    def test_column_order_strict(self):
        """Test column order validation in strict mode."""
        df = pd.DataFrame({
            "A": [1],
            "B": [2],
            "C": [3]
        })
        config = {
            "enabled": True,
            "dataframe_template": [
                {
                    "check": "column_order",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "expected_order": ["A", "B", "C"],
                        "strict": True
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise - exact match
        validator.validate_dataframe_template(df, "Test Sheet")

    def test_column_order_strict_fail(self):
        """Test strict column order validation fails with mismatch."""
        df = pd.DataFrame({
            "A": [1],
            "B": [2],
            "C": [3],
            "D": [4]  # Extra column
        })
        config = {
            "enabled": True,
            "dataframe_template": [
                {
                    "check": "column_order",
                    "enabled": True,
                    "severity": "error",
                    "params": {
                        "expected_order": ["A", "B", "C"],
                        "strict": True
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        with pytest.raises(TemplateError, match="Column order mismatch"):
            validator.validate_dataframe_template(df, "Test Sheet")

    def test_column_order_warning(self, sample_dataframe):
        """Test column order validation with warning severity."""
        config = {
            "enabled": True,
            "dataframe_template": [
                {
                    "check": "column_order",
                    "enabled": True,
                    "severity": "warning",
                    "params": {
                        "expected_order": ["Category", "Site"],
                        "strict": False
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise, just warn
        validator.validate_dataframe_template(sample_dataframe, "Test Sheet")

    def test_dataframe_validation_disabled_check(self, sample_dataframe):
        """Test disabled checks are skipped."""
        config = {
            "enabled": True,
            "dataframe_template": [
                {
                    "check": "required_columns",
                    "enabled": False,
                    "params": {
                        "columns": ["Nonexistent"]
                    }
                }
            ]
        }
        validator = TemplateValidator(config)

        # Should not raise because check is disabled
        validator.validate_dataframe_template(sample_dataframe, "Test Sheet")

    def test_dataframe_validation_disabled_validator(self, sample_dataframe):
        """Test disabled validator skips DataFrame validation."""
        config = {"enabled": False}
        validator = TemplateValidator(config)

        # Should not raise, validator is disabled
        validator.validate_dataframe_template(sample_dataframe, "Test Sheet")
